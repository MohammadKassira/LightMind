from __future__ import annotations

from pathlib import Path

_TEMPLATE_PATH = Path(__file__).parent.parent / "static" / "demand_template.xlsx"

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
PERIODS = ["Morning Rush", "Midday", "Evening Rush", "Night"]

MONDAY_PREFILL = [
    ("Morning Rush", "High",   3, "Busy commuter traffic"),
    ("Midday",       "Medium", 5, "Moderate flow"),
    ("Evening Rush", "High",   2, "Return commute"),
    ("Night",        "Low",    8, "Very quiet"),
]


def ensure_demand_template() -> None:
    """Regenerate demand_template.xlsx if it is missing or outdated."""
    if _TEMPLATE_PATH.exists():
        return
    _TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    generate_demand_template(_TEMPLATE_PATH)


def generate_demand_template(output_path: Path | None = None) -> Path:
    """Create the 7-day demand schedule template and write it to disk."""
    import openpyxl  # noqa: PLC0415
    from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415

    path = output_path or _TEMPLATE_PATH
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demand Schedule"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 30

    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    data_fill   = PatternFill(start_color="0B1929", end_color="0B1929", fill_type="solid")
    sep_fill    = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    note_font   = Font(italic=True, color="4ADE80", size=10)

    level_fills = {
        "High":   PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid"),
        "Medium": PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
        "Low":    PatternFill(start_color="86EFAC", end_color="86EFAC", fill_type="solid"),
    }
    level_fonts = {
        "High":   Font(bold=True, color="991B1B"),
        "Medium": Font(bold=True, color="92400E"),
        "Low":    Font(bold=True, color="14532D"),
    }

    # Row 1: headers
    for col, h in enumerate(["Day", "Period", "Demand_Level", "Weight", "Description"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2: note
    note = ws.cell(row=2, column=1, value="← Monday is pre-filled as an example. Fill in the rest.")
    note.font = note_font
    note.fill = data_fill
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 18

    current_row = 3
    for day_idx, day in enumerate(DAYS):
        for period_idx, period in enumerate(PERIODS):
            row = current_row
            day_cell = ws.cell(row=row, column=1, value=day if period_idx == 0 else None)
            day_cell.fill = data_fill
            day_cell.font = Font(bold=True, color="22D3C8", size=11)
            day_cell.alignment = Alignment(vertical="center")

            period_cell = ws.cell(row=row, column=2, value=period)
            period_cell.fill = data_fill
            period_cell.font = Font(color="CBD5E1", size=10)

            if day == "Monday":
                lvl, wt, desc = MONDAY_PREFILL[period_idx][1], MONDAY_PREFILL[period_idx][2], MONDAY_PREFILL[period_idx][3]
                lv_cell = ws.cell(row=row, column=3, value=lvl)
                lv_cell.fill = level_fills[lvl]
                lv_cell.font = level_fonts[lvl]
                lv_cell.alignment = Alignment(horizontal="center")
                wt_cell = ws.cell(row=row, column=4, value=wt)
                wt_cell.fill = data_fill
                wt_cell.font = Font(color="CBD5E1")
                wt_cell.alignment = Alignment(horizontal="center")
                desc_cell = ws.cell(row=row, column=5, value=desc)
                desc_cell.fill = data_fill
                desc_cell.font = Font(color="64748B", italic=True, size=10)
            else:
                for col in [3, 4, 5]:
                    ws.cell(row=row, column=col).fill = data_fill

            ws.row_dimensions[row].height = 18
            current_row += 1

        if day_idx < len(DAYS) - 1:
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).fill = sep_fill
            ws.row_dimensions[current_row].height = 8
            current_row += 1

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 60
    bg = PatternFill(start_color="0B1929", end_color="0B1929", fill_type="solid")

    def instr(row, a_val, b_val="", a_bold=False, a_color="CBD5E1", b_color="94A3B8"):
        ca = ws2.cell(row=row, column=1, value=a_val)
        ca.font = Font(bold=a_bold, color=a_color, size=10)
        ca.fill = bg
        ca.alignment = Alignment(wrap_text=True)
        if b_val:
            cb = ws2.cell(row=row, column=2, value=b_val)
            cb.font = Font(color=b_color, size=10)
            cb.fill = bg
            cb.alignment = Alignment(wrap_text=True)

    for r in ws2.iter_rows(min_row=1, max_row=30, min_col=1, max_col=3):
        for c in r:
            c.fill = bg

    title = ws2.cell(row=1, column=1, value="LightMind Demand Schedule — How to Fill")
    title.font = Font(bold=True, color="22D3C8", size=14)
    title.fill = bg
    ws2.merge_cells("A1:B1")

    instr(3,  "WHAT TO DO:",                            a_bold=True, a_color="E2E8F0")
    instr(4,  "1. Monday is pre-filled as an example",  "Use it as a reference for the format.")
    instr(5,  "2. Fill in Tuesday through Sunday.",     "Each day has 4 time periods to fill.")
    instr(6,  "3. For each period set Demand_Level",    "Must be exactly: Low, Medium, or High")
    instr(7,  "4. Set Weight",                          "Positive integer — how common is this pattern?")
    instr(8,  "5. Description is optional",             "Ignored by the system, just for your notes.")
    instr(10, "DEMAND LEVELS:",                         a_bold=True, a_color="E2E8F0")
    instr(11, "High",   "Heavy traffic — rush hour, busy periods",  a_color="FCA5A5")
    instr(12, "Medium", "Moderate traffic — normal flow",           a_color="FDE68A")
    instr(13, "Low",    "Light traffic — nights, quiet periods",    a_color="86EFAC")
    instr(15, "WEIGHT MEANING:",                        a_bold=True, a_color="E2E8F0")
    instr(16, "Higher weight = more training", "A weight of 8 vs 3 means 2.7x more episodes on that pattern.")
    instr(17, "Total weights across all rows", "Determines what % of training each period gets.")
    instr(18, "Example:",                      "Weight 3 out of total 18 = 16.7% of training episodes.")
    instr(20, "IMPORTANT:",                             a_bold=True, a_color="E2E8F0")
    instr(21, "Demand_Level is case-sensitive", "Write exactly: Low, Medium, or High")
    instr(22, "Weight must be a number",        "Use whole numbers like 1, 2, 3, 5, 8")
    instr(23, "Empty rows are ignored",         "You can leave Description blank")
    instr(24, "All 7 days should be filled",    "Unfilled rows default to Medium weight 1")

    wb.save(path)
    return path
